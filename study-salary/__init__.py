from openpyxl import load_workbook #处理excle使用
import smtplib  #邮件发送使用
from email.mime.text import MIMEText  #赋值邮件正文
from email.header import Header      #赋值邮件头
import time

#加载当前时间
now_time=time.strftime('%Y年%m月')

#加载excle文件
wb = load_workbook("salary.xlsx",data_only=True)
sheetname = wb.active

#登录邮件服务器
smtp_obj=smtplib.SMTP_SSL("smtp.qq.com",465)  #smtp.qq.com
smtp_obj.login("401360492@qq.com","qpaspyagcravbjhc")

#构建邮件正文
for row in sheetname.iter_rows(min_row=2,max_col=9):
    row_text=''
    for x in row:
        print(x.value,end=",")
        row_text += f"{x.value},"
    emp_name=row[1].value
    emp_email=row[8].value
    emp_department=row[2].value
    print()

    salary_msg=f'''
    <h3>{emp_name}您好：<h3>
    <p>请查收{now_time}工资条
    {row_text}
    </p>
    '''
    #构建头部数据
    msg = MIMEText(salary_msg,'html','utf-8')
    msg['From'] = '人事部'
    msg['To'] = '员工'
    msg['subject'] = Header('工资条','utf-8')

    #发送邮件
    smtp_obj.sendmail("401360492@qq.com",emp_email,msg.as_string())
    print(f'成功发送工资条到{emp_department}员工{emp_name} {emp_email}邮箱')





