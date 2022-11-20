# #os sys模块
# import os
# import random
# import sys
#
# print(sys.path)
# from TestP.B.BB import B1
# print(os.listdir())
# print(sys.argv)
#
# #for语法
# for i in range(10):#i是临时变量  每次循环range的值给到i  range(10) 为0-10 不包括10
#     print(i)
#
# #time模块
# import time
# start_time=time.time()
# # time.sleep(2)
# print(f"time.time()-start_time")
# print(f"{time.time()-start_time}")
#
# print(time.localtime(121212121)) #将时间戳转换为struct_time 不填时间戳则默认生成当前的struct_time
# print(time.mktime(time.localtime(121212121)))#将struct_time转化为时间戳 不填则默认生成当前时间戳
# print(time.gmtime())#将时间戳转换为struct_time(UTC 0时区) 不填时间戳则默认生成当前的UTC 0时区失去struct_time
#
#
# print(time.strftime("%Y-%m/%d")) #Y年 m月 d日 H时 M分 S秒    把一把代表时间的元组或者struct_time转化为格式化的时间字符串
# test_strptime=time.strftime("%Y-%m/%d")
# print(time.strptime(test_strptime,"%Y-%m/%d"))#字符串转时间
#
#
# #datetime模块
#
import datetime
#
# print(datetime.datetime.now())
# print(datetime.datetime.fromtimestamp(1212121))
# #时间运算
# test_datetime=datetime.datetime.now()
# print(test_datetime)
# print(test_datetime+datetime.timedelta(4,hours=1))#timedelat默认天
#
# #时间替换
# print(test_datetime.replace(year=11,month=12))
#
# #random模块
# print(random.randrange(1,10))#1-10随机数 不包括10
# print(random.randint(1,10))#1-10随机数 包括10
# print(random.random())#随机浮点数
# print(random.sample('abc123',2))#从多个字符串中选取特定数量的字符
#
# test_random=random.sample('abcdef123456',3)
# print(test_random)
# print(random.shuffle(test_random))#会返回none 因为shuffle方法不可打印
# # random.shuffle(test_random)
# print(test_random)

#EXCEl处理
from openpyxl import Workbook,load_workbook
# wb = Workbook() #在ram中打开excel文件
# print(wb.active)
# sheet =wb.active
# print(sheet.title)
# sheet.title='test_title'

#写数据 方式1 单条
# sheet['B9']="B9测试"
# sheet['A9']="A9测试"

#写数据 方式2
# sheet.append(['测试一行添加1','测试一行添加2','测试一行添加3'])#从最后一行开始添加


#写数据 加时间
# sheet.append(['测试时间添加1','测试时间添加2',datetime.datetime.now().strftime('%Y/%m/%d')])#从最后一行开始添加
# wb.save("excel_test.xlsx")


wb = load_workbook("excel_test2.xlsx")#打开文件
print(f"sheet name:{wb.sheetnames}" )
sheet = wb['test_title']#新写法 旧写法 sheet= wb.get_sheet_by_name('test_title')
#sheet =wb.active
# print(f"没有.value:{sheet['A9']}")
# print(f"有.value:{sheet['A9'].value}")


# for cell in sheet['A1:A5']: #获取切片数据
#     print(cell)
#     print(cell[0].value)

# #按照行遍历
# for row in sheet:
#     print(row)
#     for x in row:
#         print(x)
#         print(x.value)
#
# #按照列遍历
# for column in sheet.columns:
#     print(column)
#     for x in column:
#         print(x)
#         print(x.value)

#指定 行 列 遍历

for row in sheet.iter_rows(min_row=2,max_row=4,max_col=2):#从第二行 到 第四行 取前两列
    for x in row:
        print(x.value,end=',')
    print()

print()

for column in sheet.iter_cols(min_col=2,max_col=3,min_row=2,max_row=6):#从第二列 到 第三列 取第2行 到第6行
    for x in column:
        print(x.value,end=',')
    print()



# from openpyxl.styles import fonts,colors,alignment #字体类 颜色类 对齐类
#
# #python发邮件
# #python支持smtp的有 smtplib 和 email两个模块  email负责构造邮件  smtplib负责发送邮件 他对smtp协议进行了简单的封装
# import smtplib
# from email.mime.text import MIMEText #邮件正文
# from email.header import Header #邮件头
#
# #登录邮件服务器
# smtp_obj=smtplib.SMTP_SSL("smtp.qq.com",465)  #smtp.qq.com
# smtp_obj.login("401360492@qq.com","nbktuzgerqbobhha")
#
# #设置邮件头
# msg = MIMEText("这是邮件内容!",'plain','utf-8')#plain表示文本格式
# msg['From'] = '401360492@qq.com'
# msg['To'] = '3317973699@qq.com'
# msg['subject'] = Header('测试subject','utf-8')
#
# #发送
# smtp_obj.sendmail("401360492@qq.com","3317973699@qq.com",msg.as_string())#发送者，接收者，信息











